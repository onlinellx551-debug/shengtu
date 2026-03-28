from __future__ import annotations

import json
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from PIL import Image, ImageDraw, ImageEnhance, ImageFont, ImageOps


ROOT = Path(__file__).resolve().parent
STEP6_DIR = ROOT / "step6_output"
MATERIALS_ROOT = next(
    p for p in STEP6_DIR.iterdir() if p.is_dir() and not p.name.startswith(("T01_", "tmp", "oakln", "supplier"))
)
BUNDLE_DIR = MATERIALS_ROOT / "T01_bundle_2026-03-24"
ROUND1_DIR = MATERIALS_ROOT / "T01_round1_2026-03-24"
GEN_DIR = ROUND1_DIR / "01_round1_generated"
COMPARE_DIR = ROUND1_DIR / "02_reference_compare"
PROMPT_DIR = ROUND1_DIR / "03_generation_requests"
PREVIEW_DIR = ROUND1_DIR / "04_web_preview"

for folder in [ROUND1_DIR, GEN_DIR, COMPARE_DIR, PROMPT_DIR, PREVIEW_DIR]:
    folder.mkdir(parents=True, exist_ok=True)


def load_font(size: int, bold: bool = False) -> ImageFont.FreeTypeFont:
    candidates = [
        Path(r"C:\Windows\Fonts\msyhbd.ttc") if bold else Path(r"C:\Windows\Fonts\msyh.ttc"),
        Path(r"C:\Windows\Fonts\simhei.ttf"),
        Path(r"C:\Windows\Fonts\arial.ttf"),
    ]
    for path in candidates:
        if path.exists():
            try:
                return ImageFont.truetype(str(path), size=size)
            except OSError:
                continue
    return ImageFont.load_default()


FONT_H1 = load_font(42, bold=True)
FONT_H2 = load_font(28, bold=True)
FONT_BODY = load_font(22)
FONT_SMALL = load_font(18)

BG = "#f7f4ef"
CARD = "#ffffff"
TEXT = "#1d1d1f"
MUTED = "#7a7267"
LINE = "#e6ded1"
ACCENT = "#16365d"


def cover(image: Image.Image, size: tuple[int, int]) -> Image.Image:
    return ImageOps.fit(image.convert("RGB"), size, method=Image.Resampling.LANCZOS)


def contain(image: Image.Image, size: tuple[int, int]) -> Image.Image:
    return ImageOps.contain(image.convert("RGB"), size, method=Image.Resampling.LANCZOS)


def card_canvas(size: tuple[int, int] = (1080, 1080), bg: str = BG) -> Image.Image:
    return Image.new("RGB", size, bg)


def paste_center(canvas: Image.Image, img: Image.Image, xy: tuple[int, int]) -> None:
    canvas.paste(img, xy)


def draw_title(draw: ImageDraw.ImageDraw, title: str, subtitle: str) -> None:
    draw.text((70, 58), title, fill=TEXT, font=FONT_H1)
    draw.text((70, 114), subtitle, fill=MUTED, font=FONT_BODY)


def write_md(path: Path, text: str) -> None:
    path.write_text(text, encoding="utf-8")


def relative(from_dir: Path, to_path: Path) -> str:
    return to_path.relative_to(from_dir).as_posix()


pack_dir = BUNDLE_DIR / "01_upload_pack_v2"
main_dir = next(p for p in pack_dir.iterdir() if p.is_dir() and p.name.startswith("01_"))
detail_dir = next(p for p in pack_dir.iterdir() if p.is_dir() and p.name.startswith("02_"))
color_dir = next(p for p in pack_dir.iterdir() if p.is_dir() and p.name.startswith("03_"))
rec_dir = next(p for p in pack_dir.iterdir() if p.is_dir() and p.name.startswith("04_"))
ref_dir = BUNDLE_DIR / "03_reference_source_images"

main_imgs = sorted(p for p in main_dir.iterdir() if p.is_file())
detail_imgs = sorted(p for p in detail_dir.iterdir() if p.is_file())
color_imgs = sorted(p for p in color_dir.iterdir() if p.is_file())
rec_imgs = sorted(p for p in rec_dir.iterdir() if p.is_file())
ref_imgs = [p for p in sorted(ref_dir.iterdir()) if p.is_file() and p.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp", ".avif"}]
ref_map = {i + 1: p for i, p in enumerate(ref_imgs)}

# Current usable images
hero_white = main_imgs[0]
onbody = main_imgs[1]
hanger = main_imgs[2]
collar = main_imgs[3]
fold = main_imgs[4]
rack = main_imgs[5]
multi = main_imgs[6]
gray = main_imgs[7]

white = next(p for p in color_imgs if "白色" in p.name)
black = next(p for p in color_imgs if "黑色" in p.name)
green = next(p for p in color_imgs if "军绿" in p.name)
beige = next(p for p in color_imgs if "米白" in p.name)
multi_color = next(p for p in color_imgs if "多色" in p.name)


def save(img: Image.Image, filename: str) -> Path:
    path = GEN_DIR / filename
    img.save(path, quality=95)
    return path


generated_rows: list[dict[str, str]] = []


def make_g01() -> Path:
    canvas = card_canvas()
    draw = ImageDraw.Draw(canvas)
    draw_title(draw, "主图首图", "白 T 正面平铺，继续保持干净白底和留白比例")
    art = contain(Image.open(hero_white), (820, 820))
    paste_center(canvas, art, ((1080 - art.width) // 2, 220))
    return save(canvas, "G01_主图首图_白T正面平铺.jpg")


def make_stack_from_colors() -> Path:
    canvas = card_canvas()
    draw = ImageDraw.Draw(canvas)
    draw_title(draw, "面料与色组图", "用当前白 T 系列颜色重做参考页的折叠堆叠感")
    warm = Image.new("RGB", (940, 720), "#d6b397")
    canvas.paste(warm, (70, 220))
    colors = [beige, white, gray, black]
    offsets = [(120, 370), (260, 330), (400, 290), (540, 250)]
    for src, (x, y) in zip(colors, offsets):
        img = contain(Image.open(src), (340, 340))
        img = ImageEnhance.Contrast(img).enhance(1.04)
        canvas.paste(img, (x, y))
    return save(canvas, "G05_面料与色组图_折叠堆叠.jpg")


def make_detail_card(src_path: Path, filename: str, title: str, subtitle: str) -> Path:
    canvas = card_canvas()
    draw = ImageDraw.Draw(canvas)
    draw_title(draw, title, subtitle)
    art = contain(Image.open(src_path), (860, 760))
    paste_center(canvas, art, ((1080 - art.width) // 2, 220))
    return save(canvas, filename)


def make_g07_fabric_layers() -> Path:
    canvas = card_canvas()
    draw = ImageDraw.Draw(canvas)
    draw_title(draw, "布面叠放图", "用当前白 T 系列色卡模拟参考页的布面叠放镜头")
    bg = Image.new("RGB", (940, 720), "#d7b59a")
    canvas.paste(bg, (70, 220))
    layers = [white, beige, gray, black]
    offsets = [(110, 520), (210, 450), (320, 380), (430, 310)]
    for src, (x, y) in zip(layers, offsets):
        img = contain(Image.open(src), (500, 500))
        img = img.rotate(-8, expand=True, fillcolor="#d7b59a")
        canvas.paste(img, (x, y), None)
    return save(canvas, "G07_细节卖点图2_布面叠放.jpg")


def make_color_cards() -> list[Path]:
    out = []
    mapping = [
        ("G10_白色平铺色卡.jpg", white, "白色"),
        ("G10_黑色平铺色卡.jpg", black, "黑色"),
        ("G10_灰色平铺色卡.jpg", gray, "灰色"),
        ("G10_军绿平铺色卡.jpg", green, "军绿"),
        ("G10_米白平铺色卡.jpg", beige, "米白"),
    ]
    for filename, src, label in mapping:
        canvas = card_canvas()
        draw = ImageDraw.Draw(canvas)
        draw_title(draw, f"{label} 色卡", "统一白底，作为详情页后段颜色展示")
        art = contain(Image.open(src), (780, 780))
        paste_center(canvas, art, ((1080 - art.width) // 2, 220))
        out.append(save(canvas, filename))
    return out


def make_duo_card(left_src: Path, right_src: Path, filename: str, title: str) -> Path:
    canvas = card_canvas()
    draw = ImageDraw.Draw(canvas)
    draw_title(draw, title, "两件并列，作为推荐位和组合推荐卡")
    box_w = 420
    art1 = contain(Image.open(left_src), (box_w, 580))
    art2 = contain(Image.open(right_src), (box_w, 580))
    paste_center(canvas, art1, (110 + (box_w - art1.width) // 2, 280))
    paste_center(canvas, art2, (550 + (box_w - art2.width) // 2, 280))
    return save(canvas, filename)


def make_texture_micro() -> Path:
    img = Image.open(fold)
    crop = cover(img.crop((140, 190, img.width - 120, img.height - 80)), (860, 760))
    crop = ImageEnhance.Contrast(crop).enhance(1.08)
    crop = ImageEnhance.Sharpness(crop).enhance(1.15)
    return make_detail_card_from_image(crop, "G13_中段卖点图1_面料表面近景.jpg", "面料表面近景", "用现有白 T 细节图逼近参考页的柔焦近景质感")


def make_detail_card_from_image(image: Image.Image, filename: str, title: str, subtitle: str) -> Path:
    canvas = card_canvas()
    draw = ImageDraw.Draw(canvas)
    draw_title(draw, title, subtitle)
    art = contain(image, (860, 760))
    paste_center(canvas, art, ((1080 - art.width) // 2, 220))
    return save(canvas, filename)


def make_texture_pattern() -> Path:
    img = Image.open(collar)
    crop = cover(img.crop((120, 130, img.width - 120, img.height - 80)), (860, 760))
    crop = ImageEnhance.Contrast(crop).enhance(1.12)
    return make_detail_card_from_image(crop, "G14_中段卖点图2_肌理微距.jpg", "肌理微距图", "当前先用领口区域肌理替代，后续可用平台生成更像参考页的织感微距")


def add_row(gid: str, path: Path, kind: str, status: str, note: str) -> None:
    generated_rows.append(
        {
            "清单ID": gid,
            "输出文件": str(path),
            "类型": kind,
            "当前状态": status,
            "说明": note,
        }
    )


g01 = make_g01()
add_row("G01", g01, "已生成", "可直接用", "白 T 正面平铺首图，已经比较贴近参考页主图首屏逻辑")

g05 = make_stack_from_colors()
add_row("G05", g05, "已生成", "可直接用", "本地重做的折叠堆叠图，用于参考页的色组和面料气氛")

g06 = make_detail_card(collar, "G06_细节卖点图1_折边结构近景.jpg", "折边结构近景", "用现有领口细节图替代参考页的折边近景")
add_row("G06", g06, "已生成", "可直接用", "现有领口细节可先承担该位置")

g07 = make_g07_fabric_layers()
add_row("G07", g07, "已生成", "可直接用", "本地重做的布面叠放图，风格更靠近参考页")

g08 = make_detail_card(collar, "G08_细节卖点图3_领口特写.jpg", "领口特写", "重点表现圆领厚度和整洁度")
add_row("G08", g08, "已生成", "可直接用", "这张是当前最接近参考页领口特写的图")

g09 = make_detail_card(rack, "G09_细节卖点图4_袖口下摆近景.jpg", "袖口与下摆近景", "当前先用排挂近景承担局部结构展示，后续可升级")
add_row("G09", g09, "已生成", "可先用", "可用，但还不是真正对应参考页手部入镜的局部镜头")

for idx, p in enumerate(make_color_cards(), start=1):
    add_row("G10", p, "已生成", "可直接用", f"第 {idx} 张颜色平铺色卡")

g11a = make_duo_card(white, black, "G11_双拼推荐卡_白黑.jpg", "双拼推荐卡：白 + 黑")
g11b = make_duo_card(white, gray, "G11_双拼推荐卡_白灰.jpg", "双拼推荐卡：白 + 灰")
add_row("G11", g11a, "已生成", "可直接用", "白黑推荐卡")
add_row("G11", g11b, "已生成", "可直接用", "白灰推荐卡")

g13 = make_texture_micro()
add_row("G13", g13, "已生成", "可先用", "面料表面近景已经更像参考页，但还不是平台级重绘")

g14 = make_texture_pattern()
add_row("G14", g14, "已生成", "可先用", "肌理微距图先用领口区域替代")


reference_targets = {
    "G01": 2,
    "G02": 3,
    "G03": 4,
    "G05": 6,
    "G06": 7,
    "G07": 8,
    "G08": 9,
    "G09": 10,
    "G12": 39,
    "G13": 40,
    "G14": 41,
    "G15": 42,
    "G16": 43,
    "G17": 44,
}

current_map: dict[str, Path] = {
    "G01": g01,
    "G05": g05,
    "G06": g06,
    "G07": g07,
    "G08": g08,
    "G09": g09,
    "G13": g13,
    "G14": g14,
}


def compare_board(gid: str, ref_idx: int, current_path: Path | None, title: str, verdict: str, note: str) -> Path:
    canvas = Image.new("RGB", (1800, 1080), BG)
    draw = ImageDraw.Draw(canvas)
    draw.text((60, 40), f"{gid} 对比板", fill=TEXT, font=FONT_H1)
    draw.text((60, 100), title, fill=MUTED, font=FONT_BODY)

    ref_img = contain(Image.open(ref_map[ref_idx]), (780, 780))
    canvas.paste(ref_img, (80 + (780 - ref_img.width) // 2, 180))
    draw.text((80, 970), f"左：参考图 #{ref_idx}", fill=TEXT, font=FONT_H2)

    if current_path and current_path.exists():
        cur_img = contain(Image.open(current_path), (780, 780))
        canvas.paste(cur_img, (940 + (780 - cur_img.width) // 2, 180))
        draw.text((940, 970), "右：本轮重做/替代图", fill=TEXT, font=FONT_H2)
    else:
        draw.rounded_rectangle((940, 180, 1720, 960), radius=24, outline=LINE, width=3, fill=CARD)
        draw.text((1010, 390), "当前没有可直接本地生成的成品图", fill=TEXT, font=FONT_H2)
        draw.text((1010, 450), "这类镜头需要平台级生成或重绘", fill=MUTED, font=FONT_BODY)
        draw.text((1010, 510), "后续按提示词和参考图去生成", fill=MUTED, font=FONT_BODY)
        draw.text((940, 970), "右：待平台生成", fill=TEXT, font=FONT_H2)

    draw.rounded_rectangle((60, 1000, 1740, 1060), radius=16, fill=CARD, outline=LINE)
    draw.text((90, 1017), f"结论：{verdict} | {note}", fill=TEXT, font=FONT_SMALL)
    out = COMPARE_DIR / f"{gid}_compare.jpg"
    canvas.save(out, quality=95)
    return out


compare_rows: list[dict[str, str]] = []

compare_specs = [
    ("G01", "主图首图", "保留", "当前图已经能用，只需继续微调投放顺序"),
    ("G02", "上半身正面模特图", "待生成", "必须走平台生成，当前没有同模特白 T 成品图"),
    ("G03", "上半身三分之二模特图", "待生成", "必须走平台生成，当前没有同模特白 T 成品图"),
    ("G05", "折叠堆叠图", "保留", "本地重做后已经接近参考页氛围"),
    ("G06", "折边结构近景", "保留", "当前领口细节图可先承担这个镜头"),
    ("G07", "布面叠放图", "保留", "本地重做后可直接上详情"),
    ("G08", "领口特写", "保留", "当前图能直接用，后续可继续提高清晰度"),
    ("G09", "袖口/下摆近景", "可升级", "先用排挂近景替位，后续应补手部入镜局部图"),
    ("G12", "坐姿生活方式图", "待生成", "这是高优先级缺口，需要平台生成"),
    ("G13", "面料表面近景", "可升级", "当前版能用，但离参考页的高级质感还有差距"),
    ("G14", "肌理微距图", "可升级", "当前版能用，但还不是真正微距织感"),
    ("G15", "全身搭配图 1", "待生成", "需要平台生成"),
    ("G16", "全身搭配图 2", "待生成", "需要平台生成"),
    ("G17", "全身搭配图 3", "待生成", "需要平台生成"),
]

for gid, title, verdict, note in compare_specs:
    out = compare_board(gid, reference_targets[gid], current_map.get(gid), title, verdict, note)
    compare_rows.append(
        {
            "清单ID": gid,
            "参考图编号": reference_targets[gid],
            "对比图": str(out),
            "结论": verdict,
            "说明": note,
        }
    )


prompt_rows = [
    {
        "清单ID": "G02",
        "优先级": "高",
        "参考图编号": "03 / 49",
        "目标": "白 T 上半身正面模特图",
        "是否必须同模特": "是",
        "提示词": "保持参考图同一位成熟男性模特、同样的胸口以上构图、同样的室内极简背景和柔和自然光，把上衣替换为白色 300g 重磅圆领短袖 T 恤，保留干净、通勤、简洁的男装气质，不要改变模特长相、年龄感、发型、姿态和镜头焦段。",
    },
    {
        "清单ID": "G03",
        "优先级": "高",
        "参考图编号": "04 / 50",
        "目标": "白 T 上半身三分之二模特图",
        "是否必须同模特": "是",
        "提示词": "保持参考图同一位成熟男性模特和同样的站姿、裁切、背景、光线，把毛衣替换为白色圆领短袖重磅 T 恤，保持高级日系通勤感，不要改变模特面部、发型、肤色和身体比例。",
    },
    {
        "清单ID": "G12",
        "优先级": "高",
        "参考图编号": "39",
        "目标": "坐姿生活方式图",
        "是否必须同模特": "是",
        "提示词": "保持参考图里同一位男性模特、同样的坐姿、同样的暖光室内环境和笔记本入镜方式，把上衣改成白色 300g 重磅圆领短袖 T 恤，保留成熟、安静、通勤友好的日系男装风格。",
    },
    {
        "清单ID": "G15",
        "优先级": "高",
        "参考图编号": "42",
        "目标": "全身搭配图 1",
        "是否必须同模特": "是",
        "提示词": "保持参考图里同一位男性模特、同样的全身站姿、同样的浅色裤装和极简背景，把上衣换成白色重磅圆领短袖 T 恤，整体保持干净、利落、成熟的男装搭配感。",
    },
    {
        "清单ID": "G16",
        "优先级": "高",
        "参考图编号": "43",
        "目标": "全身搭配图 2",
        "是否必须同模特": "是",
        "提示词": "保持参考图里同一位男性模特、同样的全身站姿、外套层次和通勤背景，把内搭改成白色圆领重磅 T 恤，保持原有搭配比例和成熟男装气质。",
    },
    {
        "清单ID": "G17",
        "优先级": "高",
        "参考图编号": "44",
        "目标": "全身搭配图 3",
        "是否必须同模特": "是",
        "提示词": "保持参考图里同一位男性模特、同样的站姿、层次和外套搭配，把内搭改成白色重磅圆领短袖 T 恤，保留原画面的日系成熟男装风格和构图。",
    },
]

write_md(
    PROMPT_DIR / "generation_requests.md",
    "\n".join(
        [
            "# T01 平台生成请求清单",
            "",
            "这批镜头当前无法只靠本地现有同款图完成，必须走平台级生成或重绘。",
            "",
        ]
        + [
            "\n".join(
                [
                    f"## {row['清单ID']} {row['目标']}",
                    f"- 参考图编号：`{row['参考图编号']}`",
                    f"- 优先级：`{row['优先级']}`",
                    f"- 是否必须同模特：`{row['是否必须同模特']}`",
                    f"- 提示词：{row['提示词']}",
                ]
            )
            for row in prompt_rows
        ]
    ),
)


wb = Workbook()
ws = wb.active
ws.title = "generated"
ws.append(["清单ID", "输出文件", "类型", "当前状态", "说明"])
for row in generated_rows:
    ws.append([row["清单ID"], row["输出文件"], row["类型"], row["当前状态"], row["说明"]])

ws2 = wb.create_sheet("compare")
ws2.append(["清单ID", "参考图编号", "对比图", "结论", "说明"])
for row in compare_rows:
    ws2.append([row["清单ID"], row["参考图编号"], row["对比图"], row["结论"], row["说明"]])

ws3 = wb.create_sheet("prompt_requests")
ws3.append(["清单ID", "优先级", "参考图编号", "目标", "是否必须同模特", "提示词"])
for row in prompt_rows:
    ws3.append([row["清单ID"], row["优先级"], row["参考图编号"], row["目标"], row["是否必须同模特"], row["提示词"]])

workbook_path = ROUND1_DIR / "T01_round1_compare_2026-03-24.xlsx"
wb.save(workbook_path)


def preview_html(paths: Iterable[Path], title: str) -> str:
    blocks = []
    for path in paths:
        blocks.append(
            f'<div class="card"><img src="../01_round1_generated/{path.name}" alt=""><div class="cap">{path.stem}</div></div>'
        )
    return f"<section><h2>{title}</h2><div class='grid'>{''.join(blocks)}</div></section>"


generated_paths = sorted(GEN_DIR.glob("*.jpg"))
html = f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>T01 Round1 对比预览</title>
  <style>
    body {{ margin: 0; font-family: Arial, "Microsoft YaHei", sans-serif; background: #f7f4ef; color: #1d1d1f; }}
    .wrap {{ max-width: 1380px; margin: 0 auto; padding: 30px; }}
    h1 {{ font-size: 34px; margin: 0 0 10px; }}
    p {{ color: #6f675e; line-height: 1.7; }}
    section {{ margin-top: 36px; }}
    h2 {{ font-size: 22px; margin-bottom: 14px; }}
    .grid {{ display: grid; grid-template-columns: repeat(3, minmax(0, 1fr)); gap: 16px; }}
    .card {{ background: white; border: 1px solid #eadfce; padding: 12px; }}
    img {{ width: 100%; display: block; }}
    .cap {{ font-size: 12px; color: #6f675e; margin-top: 8px; word-break: break-all; }}
  </style>
</head>
<body>
  <div class="wrap">
    <h1>T01 Round1 本地重做与对比预览</h1>
    <p>这一轮只重做当前能本地优化的镜头，并把必须走平台生成的镜头单独拆出来。高优先级缺口仍然是同模特上身图和全身搭配图。</p>
    {preview_html(generated_paths, "本轮可用新素材")}
  </div>
</body>
</html>
"""
(PREVIEW_DIR / "index.html").write_text(html, encoding="utf-8")

summary = {
    "generated_count": len(generated_rows),
    "compare_count": len(compare_rows),
    "prompt_request_count": len(prompt_rows),
    "round1_dir": str(ROUND1_DIR),
    "preview_html": str(PREVIEW_DIR / "index.html"),
    "workbook": str(workbook_path),
}
(ROUND1_DIR / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

write_md(
    ROUND1_DIR / "README.md",
    "\n".join(
        [
            "# T01 Round1 结果",
            "",
            f"- 生成图目录：`{GEN_DIR}`",
            f"- 对比图目录：`{COMPARE_DIR}`",
            f"- 平台生成请求：`{PROMPT_DIR / 'generation_requests.md'}`",
            f"- 工作簿：`{workbook_path}`",
            f"- 网页预览：`{PREVIEW_DIR / 'index.html'}`",
            "",
            "说明：",
            "- 这一轮先把可以本地优化的镜头重做出来。",
            "- 同模特上身图、坐姿图、全身搭配图仍然缺平台级生成。",
            "- 下一步如果继续，就应该直接执行 G02 / G03 / G12 / G15 / G16 / G17。",
        ]
    ),
)

print(workbook_path)
print(PREVIEW_DIR / "index.html")
